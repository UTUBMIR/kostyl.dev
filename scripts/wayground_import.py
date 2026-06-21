#!/usr/bin/env python3
"""
wayground_import.py — Автоматичний імпорт xlsx-тесту в Wayground.

Pipeline:
  1. Логін → отримати _sid cookie
  2. Завантажити .xlsx на S3 через Wayground media API
  3. POST /upload-quiz → розпарсити питання з xlsx
  4. POST /v3/quiz → створити порожній тест, отримати quiz_id + version_id
  5. POST /questions → додати всі питання у тест
  6. Вивести посилання на готовий тест

Використання:
  python3 wayground_import.py <шлях_до_xlsx>
  python3 wayground_import.py <шлях_до_xlsx> --name "Назва тесту"
  python3 wayground_import.py <шлях_до_xlsx> --name "Назва" --lang uk
  python3 wayground_import.py <шлях_до_xlsx> --batch-size 5   # питань за один POST

Credentials — з .env поряд зі скриптом або зі змінних середовища:
  WAYGROUND_EMAIL=...
  WAYGROUND_PASSWORD=...
"""

import os
import sys
import uuid
import json
import argparse
import requests
from pathlib import Path
from urllib.parse import quote


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

QUESTION_THEME = {
    "titleFontFamily": "Quicksand",
    "fontFamily": "Quicksand",
    "fontColor": {"text": "#5D2057"},
    "background": {"color": "#FFFFFF", "image": "", "video": ""},
    "shape": {"largeShapeColor": "#E9E0F3", "smallShapeColor": "#9A4292"},
}


# ---------------------------------------------------------------------------
# Auth (shared з wayground_add_images.py)
# ---------------------------------------------------------------------------

def load_credentials() -> tuple[str, str]:
    for env_dir in [Path(__file__).parent, Path.cwd()]:
        env_file = env_dir / ".env"
        if env_file.exists():
            for line in env_file.read_text().splitlines():
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, _, val = line.partition("=")
                    os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))
            break
    return os.environ.get("WAYGROUND_EMAIL", ""), os.environ.get("WAYGROUND_PASSWORD", "")


def login(email: str, password: str) -> requests.Session:
    print(f"[~] Авторизація як {email}...")
    resp = requests.post(
        "https://wayground.com/_authserver/public/public/v1/auth/login/local",
        json={"username": email, "password": password, "requestId": str(uuid.uuid4())},
        headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Origin": "https://wayground.com",
        },
        timeout=30,
    )
    if resp.status_code != 200 or not resp.json().get("success"):
        raise RuntimeError(f"Помилка авторизації. HTTP {resp.status_code}: {resp.text[:200]}")

    sid = resp.cookies.get("_sid")
    quizizz_uid = resp.cookies.get("quizizz_uid")
    if not sid:
        raise RuntimeError("_sid cookie не знайдено у відповіді login")

    session = requests.Session()
    session.cookies.set("_sid", sid, domain="wayground.com")
    if quizizz_uid:
        session.cookies.set("quizizz_uid", quizizz_uid, domain="wayground.com")
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36",
        "Origin": "https://wayground.com",
        "Content-Type": "application/json",
    })

    user = resp.json().get("data", {}).get("user", {})
    name = f"{user.get('firstName', '')} {user.get('lastName', '')}".strip()
    print(f"[✓] Авторизовано: {name}")
    return session


# ---------------------------------------------------------------------------
# Крок 1: Завантаження xlsx на S3
# ---------------------------------------------------------------------------

def upload_xlsx(session: requests.Session, xlsx_path: Path) -> list[dict]:
    """
    Завантажує xlsx на S3, парсить питання через Wayground API,
    потім об'єднує результат з даними з xlsx (Image Link, Answer explanation).

    Wayground /upload-quiz ігнорує колонки Image Link і Answer explanation —
    тому їх треба читати напряму з xlsx і вставляти в extracted вручну.
    """
    import openpyxl

    xlsx_data = xlsx_path.read_bytes()
    print(f"[~] Завантажуємо {xlsx_path.name} ({len(xlsx_data):,} байт) на S3...")

    # Читаємо Image Link та Answer explanation напряму з xlsx
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    xlsx_rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        xlsx_rows.append(row_dict)

    # Отримуємо pre-signed URL
    metadata = quote(f'{{"Content-Type":"{XLSX_CONTENT_TYPE}"}}')
    init_url = (
        "https://media.quizizz.com/_mdserver/main/getUploadURL"
        f"?destination=uploadedSheets&folder=_excel&metadata={metadata}&enableAcceleration=true"
    )
    r = session.post(init_url, timeout=30)
    if r.status_code != 200 or not r.json().get("success"):
        raise RuntimeError(f"getUploadURL failed: {r.status_code} {r.text[:200]}")

    data = r.json()["data"]
    signed_url: str = data["signedUrl"]
    final_url: str = data["finalUrl"]
    key = "_excel/uploadedSheets/" + final_url.split("uploadedSheets/")[1]

    s3_resp = requests.put(signed_url, headers={"Content-Type": XLSX_CONTENT_TYPE}, data=xlsx_data, timeout=60)
    if s3_resp.status_code != 200:
        raise RuntimeError(f"S3 upload failed: {s3_resp.status_code} {s3_resp.text[:200]}")
    print(f"[✓] Файл завантажено на S3")

    print("[~] Парсимо питання з xlsx...")
    parse_resp = session.post(
        "https://wayground.com/_api/main/upload-quiz",
        json={"key": key},
        timeout=30,
    )
    if parse_resp.status_code != 200 or not parse_resp.json().get("success"):
        raise RuntimeError(f"upload-quiz failed: {parse_resp.status_code} {parse_resp.text[:300]}")

    extracted: list[dict] = parse_resp.json()["data"]["extracted"]

    # Вставляємо Image Link і Answer explanation з xlsx (API їх ігнорує)
    images_count = 0
    for i, e in enumerate(extracted):
        if i >= len(xlsx_rows):
            break
        row = xlsx_rows[i]
        img_link = row.get("Image Link") or ""
        explain = row.get("Answer explanation") or e.get("explain", "")
        if img_link and str(img_link).strip().startswith("http"):
            e["url"] = str(img_link).strip()
            e["mediaType"] = "image"
            images_count += 1
        if explain:
            e["explain"] = str(explain).strip()

    img_note = f", {images_count} з зображеннями" if images_count else ""
    print(f"[✓] Розпарсовано {len(extracted)} питань{img_note}")
    return extracted


# ---------------------------------------------------------------------------
# Крок 2: Створення порожнього тесту
# ---------------------------------------------------------------------------

def create_quiz(session: requests.Session, name: str, lang: str = "uk") -> tuple[str, str]:
    """
    Створює порожній quiz. Повертає (quiz_id, version_id).
    """
    print(f"[~] Створюємо тест '{name}'...")
    r = session.post(
        "https://wayground.com/_quizserver/main/v3/quiz",
        json={"name": name, "type": "quiz", "lang": lang},
        headers={"Referer": "https://wayground.com/admin/quiz/create"},
        timeout=30,
    )
    if r.status_code not in (200, 201) or not r.json().get("success"):
        raise RuntimeError(f"create quiz failed: {r.status_code} {r.text[:200]}")

    quiz = r.json()["data"]["quiz"]
    quiz_id: str = quiz["_id"]
    version_id: str = quiz["draft"]["_id"]
    print(f"[✓] Тест створено: quiz_id={quiz_id}")
    return quiz_id, version_id


# ---------------------------------------------------------------------------
# Крок 3: Конвертація extracted → API question format
# ---------------------------------------------------------------------------

def _make_option_id() -> str:
    """24-символьний hex id для варіанта відповіді."""
    return uuid.uuid4().hex[:24]


def build_question(extracted: dict) -> dict:
    """
    Конвертує одне extracted питання у формат POST /questions.

    extracted поля:
      text, questionType (MCQ/MSQ), option1-5, correct (int 1-based або list 0-based),
      time (seconds), explain, url?, mediaType?
    """
    kind: str = extracted.get("questionType", "MCQ")
    text: str = extracted.get("text", "")
    explain_text: str = extracted.get("explain", "")
    time_sec: int = extracted.get("time", 30)

    # --- Варіанти відповідей ---
    options = []
    for i in range(1, 6):
        opt_text = extracted.get(f"option{i}")
        if opt_text is not None:
            options.append({
                "text": str(opt_text),
                "_id": _make_option_id(),
                "math": {"latex": [], "template": None},
                "type": "text",
                "hasMath": False,
                "media": [],
            })

    # --- Правильна відповідь ---
    correct_raw = extracted.get("correct")
    if kind == "MSQ":
        # extracted: 0-based list [0, 1, 2] → конвертуємо в 1-based для API
        raw_list = correct_raw if isinstance(correct_raw, list) else [correct_raw]
        answer = [x + 1 for x in raw_list]
    else:
        # extracted: 0-based int — API теж очікує 0-based
        if isinstance(correct_raw, list):
            answer = correct_raw[0] if correct_raw else 0
        else:
            answer = correct_raw if correct_raw is not None else 0

    # --- Query media (зображення до питання, якщо є) ---
    query_media = []
    img_url = extracted.get("url")
    if img_url and extracted.get("mediaType") == "image":
        query_media.append({
            "url": img_url,
            "type": "image",
            "meta": {
                "width": 0, "height": 0, "text": "", "bgColor": "",
                "layout": "", "videoId": "", "start": 0, "end": 0,
                "duration": 0, "kind": "", "embeddable": True, "title": "",
            },
        })

    query_type = "text_image" if query_media else "text"

    return {
        "time": time_sec * 1000,
        "type": kind,
        "structure": {
            "answer": answer,
            "kind": kind,
            "options": options,
            "settings": {
                "hasCorrectAnswer": True,
                "fibDataType": "string",
                "canSubmitCustomResponse": False,
                "doesOptionHaveMultipleTargets": False,
                "showAdvancedRTE": False,
            },
            "query": {
                "type": query_type,
                "text": text,
                "media": query_media,
                "answerTime": -1,
                "math": {"latex": [], "template": None},
                "hasMath": False,
            },
            "explain": {
                "text": explain_text,
                "math": {"latex": [], "template": None},
                "type": "",
                "hasMath": False,
                "media": [],
            },
            "theme": QUESTION_THEME,
            "graphs": [],
            "hints": [],
            "order": "asc",
            "hasMath": False,
            "queries": [],
            "marks": {"correct": 1, "incorrect": 0},
            "media": {},
            "elements": [],
        },
        "index": -1,
        "_id": "",
        "startedAt": "",
        "v": 0,
        "clones": [],
        "published": False,
        "deleted": False,
        "isSuperParent": False,
        "metaData": {},
        "topics": [],
        "standards": [],
        "marksUpdated": False,
        "state": "",
        "metadata": {},
    }


# ---------------------------------------------------------------------------
# Крок 4: Додавання питань до тесту
# ---------------------------------------------------------------------------

def import_questions(
    session: requests.Session,
    quiz_id: str,
    version_id: str,
    extracted_questions: list[dict],
    batch_size: int = 10,
) -> int:
    """
    Конвертує та додає питання до тесту батчами.
    Повертає кількість успішно доданих питань.
    """
    url = (
        f"https://wayground.com/_quizserver/main/v3/quiz/"
        f"{quiz_id}/version/{version_id}/questions"
    )
    referer = f"https://wayground.com/admin/quiz/{quiz_id}/edit"

    total = len(extracted_questions)
    added = 0

    for batch_start in range(0, total, batch_size):
        batch = extracted_questions[batch_start : batch_start + batch_size]
        batch_end = batch_start + len(batch)

        print(f"[~] Додаємо питання {batch_start + 1}–{batch_end} з {total}...")

        questions = [build_question(e) for e in batch]

        r = session.post(
            url,
            json={"questions": questions, "index": batch_start, "aiMeta": {}},
            headers={"Referer": referer},
            timeout=60,
        )

        if r.status_code in (200, 201) and r.json().get("success"):
            count = len(r.json().get("data", {}).get("questions", []))
            added += count
            print(f"  [✓] Додано {count} питань")
        else:
            print(f"  [X] Помилка батчу {batch_start}–{batch_end}. HTTP {r.status_code}: {r.text[:300]}")

    return added


# ---------------------------------------------------------------------------
# Крок 5: Публікація тесту
# ---------------------------------------------------------------------------

# Мова для PUT metadata: Wayground очікує повну назву мови
LANG_MAP = {
    "uk": "Ukrainian",
    "en": "English",
    "de": "German",
    "pl": "Polish",
    "fr": "French",
    "es": "Spanish",
}


def publish_quiz(
    session: requests.Session,
    quiz_id: str,
    version_id: str,
    name: str,
    lang: str = "uk",
) -> bool:
    """
    Публікує тест у два кроки:
      1. PUT /v2/quiz/{id}/version/{vid} — встановлює метадані і visibility: true
      2. POST /v3/quiz/{id}/version/{vid}/publish — фіксує опублікований стан

    Повертає True якщо обидва кроки успішні.
    """
    referer = f"https://wayground.com/admin/quiz/{quiz_id}/edit"
    lang_full = LANG_MAP.get(lang, "Ukrainian")

    # Крок 1: встановити метадані версії
    print("[~] Встановлюємо метадані тесту...")
    put_url = f"https://wayground.com/_quizserver/main/v2/quiz/{quiz_id}/version/{version_id}"
    r = session.put(
        put_url,
        json={
            "grade": ["14", "14"],
            "image": "",
            "lang": lang_full,
            "name": name,
            "subjects": ["Information Technology (IT)"],
            "visibility": True,
            "quizMetadata": {"teachingGoal": "review"},
        },
        headers={"Referer": referer},
        timeout=30,
    )
    if r.status_code != 200 or not r.json().get("success"):
        print(f"  [X] PUT version failed: {r.status_code} {r.text[:200]}")
        return False

    # Крок 2: опублікувати
    print("[~] Публікуємо тест...")
    pub_url = f"https://wayground.com/_quizserver/main/v3/quiz/{quiz_id}/version/{version_id}/publish"
    r = session.post(pub_url, json={}, headers={"Referer": referer}, timeout=30)
    if r.status_code not in (200, 201) or not r.json().get("success"):
        print(f"  [X] publish failed: {r.status_code} {r.text[:200]}")
        return False

    print("[✓] Тест опубліковано!")
    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Автоматичний імпорт xlsx-тесту в Wayground.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Приклади:
  python3 wayground_import.py tests/my_test.xlsx
  python3 wayground_import.py tests/my_test.xlsx --name "Python ООП"
  python3 wayground_import.py tests/my_test.xlsx --name "Python ООП" --lang uk --batch-size 5
  python3 wayground_import.py tests/my_test.xlsx --name "Python ООП" --no-publish

Credentials — з .env або змінних середовища:
  WAYGROUND_EMAIL=...
  WAYGROUND_PASSWORD=...
        """,
    )
    parser.add_argument("xlsx", help="Шлях до .xlsx файлу")
    parser.add_argument(
        "--name",
        help="Назва тесту (за замовчуванням — ім'я файлу без розширення)",
    )
    parser.add_argument(
        "--lang",
        default="uk",
        help="Мова тесту (uk, en, ...). За замовчуванням: uk",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=10,
        metavar="N",
        help="Кількість питань за один POST запит (за замовчуванням: 10)",
    )
    parser.add_argument(
        "--no-publish",
        action="store_true",
        help="Не публікувати тест після імпорту (залишити як чернетку)",
    )
    parser.add_argument("--email", help="Email (або WAYGROUND_EMAIL у .env)")
    parser.add_argument("--password", help="Пароль (або WAYGROUND_PASSWORD у .env)")

    args = parser.parse_args()

    # Файл
    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.exists():
        print(f"[X] Файл не знайдено: {xlsx_path}")
        sys.exit(1)
    if xlsx_path.suffix.lower() != ".xlsx":
        print(f"[X] Очікується .xlsx файл, отримано: {xlsx_path.suffix}")
        sys.exit(1)

    quiz_name = args.name or xlsx_path.stem

    # Credentials
    env_email, env_password = load_credentials()
    email = args.email or env_email
    password = args.password or env_password
    if not email or not password:
        print("[X] Credentials не знайдено. Створи .env або передай --email / --password")
        sys.exit(1)

    print(f"[~] Файл:  {xlsx_path}")
    print(f"[~] Назва: {quiz_name}")
    print(f"[~] Мова:  {args.lang}")
    print()

    try:
        # Auth
        session = login(email, password)
        print()

        # 1. Upload + parse xlsx
        extracted = upload_xlsx(session, xlsx_path)
        print()

        # 2. Створити тест
        quiz_id, version_id = create_quiz(session, quiz_name, args.lang)
        print()

        # 3. Додати питання
        print(f"[~] Імпортуємо {len(extracted)} питань (batch_size={args.batch_size})...")
        added = import_questions(session, quiz_id, version_id, extracted, args.batch_size)
        print()

        # 4. Публікація
        published = False
        if not args.no_publish:
            published = publish_quiz(session, quiz_id, version_id, quiz_name, args.lang)
            print()

        # Підсумок
        print("=" * 55)
        print(f"Готово! Додано {added} з {len(extracted)} питань.")
        if published:
            print(f"Статус:   опубліковано ✓")
        else:
            print(f"Статус:   чернетка (запусти з --no-publish щоб лишити так)")
        print(f"Редактор: https://wayground.com/admin/quiz/{quiz_id}/edit")
        print(f"quiz_id:  {quiz_id}")
        print("=" * 55)

    except RuntimeError as e:
        print(f"\n[X] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
